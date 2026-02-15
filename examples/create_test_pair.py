#!/usr/bin/env python3
"""Generate test_old.xlsx and test_new.xlsx for xlsx-review diff testing."""

import openpyxl

def create_old():
    wb = openpyxl.Workbook()

    # Sheet 1: Data (patient data, 5 rows)
    ws = wb.active
    ws.title = "Data"
    ws.append(["Patient ID", "Age", "Score", "Group"])
    ws.append(["P001", 34, 88, "Control"])
    ws.append(["P002", 28, 92, "Treatment"])
    ws.append(["P003", 45, 76, "Control"])
    ws.append(["P004", 31, 85, "Treatment"])
    ws.append(["P005", 39, 90, "Control"])

    # Sheet 2: Summary (with formulas)
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Metric"
    ws2["B1"] = "Value"
    ws2["A2"] = "Count"
    ws2["B2"] = 5
    ws2["A3"] = "Mean Age"
    ws2["B3"] = "=AVERAGE(Data!B2:B6)"
    ws2["A4"] = "Mean Score"
    ws2["B4"] = "=AVERAGE(Data!C2:C6)"
    ws2["A5"] = "Total Score"
    ws2["B5"] = "=SUM(Data!C2:C6)"

    wb.save("test_old.xlsx")
    print("Created test_old.xlsx")

def create_new():
    wb = openpyxl.Workbook()

    # Sheet 1: Data — modified values, added a row
    ws = wb.active
    ws.title = "Data"
    ws.append(["Patient ID", "Age", "Score", "Group"])
    ws.append(["P001", 34, 88, "Control"])
    ws.append(["P002", 29, 95, "Treatment"])   # Age 28→29, Score 92→95
    ws.append(["P003", 45, 76, "Control"])
    ws.append(["P004", 31, 85, "Treatment"])
    ws.append(["P005", 39, 90, "Control"])
    ws.append(["P006", 42, 81, "Treatment"])    # New row added

    # Sheet 2: Summary — changed a formula
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Metric"
    ws2["B1"] = "Value"
    ws2["A2"] = "Count"
    ws2["B2"] = 6                               # Updated count
    ws2["A3"] = "Mean Age"
    ws2["B3"] = "=AVERAGE(Data!B2:B7)"          # Updated range
    ws2["A4"] = "Median Score"                   # Changed from Mean Score
    ws2["B4"] = "=MEDIAN(Data!C2:C7)"           # Changed formula
    ws2["A5"] = "Total Score"
    ws2["B5"] = "=SUM(Data!C2:C7)"             # Updated range

    # Sheet 3: Notes (new sheet)
    ws3 = wb.create_sheet("Notes")
    ws3["A1"] = "Date"
    ws3["B1"] = "Note"
    ws3["A2"] = "2026-02-14"
    ws3["B2"] = "Added patient P006, updated summary formulas"

    wb.save("test_new.xlsx")
    print("Created test_new.xlsx")

if __name__ == "__main__":
    create_old()
    create_new()
    print("Done! Run:")
    print("  xlsx-review --diff test_old.xlsx test_new.xlsx")
    print("  xlsx-review --textconv test_new.xlsx")
